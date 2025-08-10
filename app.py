import streamlit as st
import json
import os
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
import google.generativeai as genai
import tempfile
import logging
from io import BytesIO
import pandas as pd
import math
import re
from collections import Counter, defaultdict

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Page config
st.set_page_config(
    page_title="📊 Excel AI Analyzer with Dynamic Retrieval",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

class SimpleTFIDFSearchEngine:
    """Lightweight TF-IDF search engine without sklearn dependency"""
    
    def __init__(self):
        self.documents = []
        self.row_metadata = []
        self.vocabulary = {}
        self.idf_scores = {}
        self.tfidf_vectors = []
        
    def _tokenize(self, text: str) -> List[str]:
        """Simple tokenization"""
        # Convert to lowercase, remove special chars, split on whitespace
        text = re.sub(r'[^\w\s]', ' ', text.lower())
        tokens = text.split()
        # Remove common stop words
        stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by', 'is', 'are', 'was', 'were', 'be', 'been', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should'}
        return [token for token in tokens if token not in stop_words and len(token) > 1]
    
    def _compute_tf(self, tokens: List[str]) -> Dict[str, float]:
        """Compute term frequency"""
        tf = {}
        total_tokens = len(tokens)
        token_counts = Counter(tokens)
        
        for token, count in token_counts.items():
            tf[token] = count / total_tokens
        
        return tf
    
    def _compute_idf(self) -> None:
        """Compute inverse document frequency for all terms"""
        doc_count = len(self.documents)
        term_doc_count = defaultdict(int)
        
        # Count documents containing each term
        for doc in self.documents:
            tokens = self._tokenize(doc)
            unique_tokens = set(tokens)
            for token in unique_tokens:
                term_doc_count[token] += 1
        
        # Compute IDF scores
        for term, count in term_doc_count.items():
            self.idf_scores[term] = math.log(doc_count / count)
    
    def _compute_tfidf_vector(self, tokens: List[str]) -> Dict[str, float]:
        """Compute TF-IDF vector for a document"""
        tf_scores = self._compute_tf(tokens)
        tfidf_vector = {}
        
        for token in tf_scores:
            if token in self.idf_scores:
                tfidf_vector[token] = tf_scores[token] * self.idf_scores[token]
        
        return tfidf_vector
    
    def _cosine_similarity(self, vec1: Dict[str, float], vec2: Dict[str, float]) -> float:
        """Compute cosine similarity between two TF-IDF vectors"""
        # Get common terms
        common_terms = set(vec1.keys()) & set(vec2.keys())
        
        if not common_terms:
            return 0.0
        
        # Compute dot product
        dot_product = sum(vec1[term] * vec2[term] for term in common_terms)
        
        # Compute magnitudes
        mag1 = math.sqrt(sum(val**2 for val in vec1.values()))
        mag2 = math.sqrt(sum(val**2 for val in vec2.values()))
        
        if mag1 == 0 or mag2 == 0:
            return 0.0
        
        return dot_product / (mag1 * mag2)
    
    def build_index(self, files_data: List[Dict[str, Any]]) -> None:
        """Build TF-IDF index from all Excel data"""
        self.documents = []
        self.row_metadata = []
        
        for file_data in files_data:
            if 'error' in file_data:
                continue
                
            file_name = file_data['file_name']
            
            for sheet_name, sheet_data in file_data['sheets'].items():
                if 'error' in sheet_data:
                    continue
                
                # Process each row as a document
                for row_idx, row in enumerate(sheet_data['data']):
                    # Combine all cell values into a single text document
                    text_content = []
                    for header, value in row.items():
                        if value and str(value).strip() and not header.startswith('_'):
                            text_content.append(f"{header}: {str(value)}")
                    
                    if text_content:
                        document = " | ".join(text_content)
                        self.documents.append(document)
                        
                        # Store metadata for each row
                        self.row_metadata.append({
                            'file_name': file_name,
                            'sheet_name': sheet_name,
                            'row_index': row_idx,
                            'row_data': row,
                            'headers': list(row.keys())
                        })
        
        if self.documents:
            # Compute IDF scores
            self._compute_idf()
            
            # Compute TF-IDF vectors for all documents
            self.tfidf_vectors = []
            for doc in self.documents:
                tokens = self._tokenize(doc)
                tfidf_vector = self._compute_tfidf_vector(tokens)
                self.tfidf_vectors.append(tfidf_vector)
    
    def search(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """Search for relevant rows using TF-IDF similarity"""
        if not self.documents:
            return []
        
        # Compute TF-IDF vector for query
        query_tokens = self._tokenize(query)
        query_vector = self._compute_tfidf_vector(query_tokens)
        
        if not query_vector:
            return []
        
        # Calculate similarities
        similarities = []
        for i, doc_vector in enumerate(self.tfidf_vectors):
            similarity = self._cosine_similarity(query_vector, doc_vector)
            similarities.append((i, similarity))
        
        # Sort by similarity score (descending)
        similarities.sort(key=lambda x: x[1], reverse=True)
        
        # Get top-k results
        results = []
        for idx, similarity_score in similarities[:top_k]:
            if similarity_score > 0:  # Only include relevant results
                result = self.row_metadata[idx].copy()
                result['similarity_score'] = similarity_score
                result['matched_text'] = self.documents[idx]
                results.append(result)
        
        return results

class DataRetriever:
    """Advanced data retrieval system for LLM requests"""
    
    def __init__(self, excel_processor: 'ExcelProcessor'):
        self.excel_processor = excel_processor
        self.search_engine = excel_processor.search_engine
        
    def get_specific_row(self, file_name: str, sheet_name: str, row_number: int) -> Dict[str, Any]:
        """Get a specific row by number"""
        try:
            for metadata in self.search_engine.row_metadata:
                if (metadata['file_name'] == file_name and 
                    metadata['sheet_name'] == sheet_name and 
                    metadata['row_data'].get('_row_number') == row_number):
                    return {
                        'success': True,
                        'data': metadata['row_data'],
                        'location': f"File: {file_name}, Sheet: {sheet_name}, Row: {row_number}"
                    }
            return {'success': False, 'error': f"Row {row_number} not found in {sheet_name}"}
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def get_column_values(self, file_name: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        """Get all values from a specific column"""
        try:
            values = []
            for metadata in self.search_engine.row_metadata:
                if (metadata['file_name'] == file_name and 
                    metadata['sheet_name'] == sheet_name):
                    row_data = metadata['row_data']
                    if column_name in row_data and row_data[column_name]:
                        values.append({
                            'row_number': row_data.get('_row_number', '?'),
                            'value': row_data[column_name]
                        })
            
            return {
                'success': True,
                'data': values,
                'count': len(values),
                'location': f"File: {file_name}, Sheet: {sheet_name}, Column: {column_name}"
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def get_sheet_summary(self, file_name: str, sheet_name: str) -> Dict[str, Any]:
        """Get complete summary of a sheet"""
        try:
            for file_data in st.session_state.files_data:
                if file_data['file_name'] == file_name:
                    if sheet_name in file_data['sheets']:
                        sheet_data = file_data['sheets'][sheet_name]
                        return {
                            'success': True,
                            'summary': sheet_data['summary'],
                            'sample_data': sheet_data['data'][:10],  # First 10 rows
                            'location': f"File: {file_name}, Sheet: {sheet_name}"
                        }
            return {'success': False, 'error': f"Sheet {sheet_name} not found in {file_name}"}
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def calculate_column_statistics(self, file_name: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        """Calculate statistics for a numeric column"""
        try:
            values = []
            for metadata in self.search_engine.row_metadata:
                if (metadata['file_name'] == file_name and 
                    metadata['sheet_name'] == sheet_name):
                    row_data = metadata['row_data']
                    if column_name in row_data and row_data[column_name]:
                        try:
                            # Try to convert to float
                            val = float(str(row_data[column_name]).replace(',', ''))
                            values.append(val)
                        except:
                            continue
            
            if not values:
                return {'success': False, 'error': f"No numeric values found in column {column_name}"}
            
            stats = {
                'count': len(values),
                'sum': sum(values),
                'average': sum(values) / len(values),
                'min': min(values),
                'max': max(values),
                'median': sorted(values)[len(values)//2]
            }
            
            return {
                'success': True,
                'statistics': stats,
                'location': f"File: {file_name}, Sheet: {sheet_name}, Column: {column_name}"
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    def search_and_aggregate(self, query: str, operation: str = 'list') -> Dict[str, Any]:
        """Search for data and perform aggregation"""
        try:
            search_results = self.search_engine.search(query, top_k=50)
            
            if not search_results:
                return {'success': False, 'error': f"No results found for query: {query}"}
            
            if operation == 'list':
                return {
                    'success': True,
                    'results': search_results[:10],  # Top 10 results
                    'total_found': len(search_results)
                }
            elif operation == 'sum':
                # Try to sum numeric values from search results
                total = 0
                count = 0
                for result in search_results:
                    for key, value in result['row_data'].items():
                        if not key.startswith('_'):
                            try:
                                num_val = float(str(value).replace(',', ''))
                                total += num_val
                                count += 1
                            except:
                                continue
                
                return {
                    'success': True,
                    'sum': total,
                    'values_counted': count,
                    'query': query
                }
            elif operation == 'count':
                return {
                    'success': True,
                    'count': len(search_results),
                    'query': query
                }
                
        except Exception as e:
            return {'success': False, 'error': str(e)}

class ExcelProcessor:
    """Enhanced Excel file processor with TF-IDF capabilities"""
    
    def __init__(self):
        self.max_rows_per_sheet = 100
        self.max_chars_per_cell = 500
        self.search_engine = SimpleTFIDFSearchEngine()
        
    def read_excel_file(self, file_content: bytes, file_name: str) -> Dict[str, Any]:
        """Read Excel file from bytes content"""
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(file_content)
                tmp_path = tmp_file.name
            
            workbook = load_workbook(tmp_path, data_only=True)
            
            file_data = {
                'file_name': file_name,
                'sheets': {},
                'summary': {
                    'total_sheets': len(workbook.sheetnames),
                    'sheet_names': workbook.sheetnames
                }
            }
            
            for sheet_name in workbook.sheetnames:
                sheet_data = self._process_sheet(workbook[sheet_name], sheet_name)
                file_data['sheets'][sheet_name] = sheet_data
            
            try:
                os.unlink(tmp_path)
            except:
                pass
                
            return file_data
            
        except Exception as e:
            logger.error(f"Error reading Excel file {file_name}: {str(e)}")
            return {'error': f"Failed to read {file_name}: {str(e)}"}
    
    def _process_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Process individual sheet with enhanced data capture"""
        try:
            if sheet.max_row is None or sheet.max_row == 0:
                return {
                    'sheet_name': sheet_name,
                    'summary': {'total_rows': 0, 'total_columns': 0, 'headers': [], 'sample_data': []},
                    'data': []
                }
            
            max_row = min(sheet.max_row, self.max_rows_per_sheet + 1)
            max_col = sheet.max_column or 0
            
            # Extract headers with better cleaning
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value is not None:
                    header = str(cell_value).strip()
                    # Clean header name
                    header = re.sub(r'[^\w\s]', '', header)
                    header = re.sub(r'\s+', ' ', header)
                    headers.append(header if header else f"Column_{col}")
                else:
                    headers.append(f"Column_{col}")
            
            # Extract data rows with better preservation
            data_rows = []
            for row in range(2, max_row + 1):
                row_data = {}
                has_data = False
                
                for col, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    if cell_value is not None:
                        cell_str = str(cell_value).strip()
                        # Preserve more content but limit for token efficiency
                        if len(cell_str) > self.max_chars_per_cell:
                            cell_str = cell_str[:self.max_chars_per_cell] + "..."
                        row_data[header] = cell_str
                        has_data = True
                    else:
                        row_data[header] = ""
                
                if has_data:
                    # Add row identifier
                    row_data['_row_number'] = row - 1  # 1-based numbering
                    data_rows.append(row_data)
            
            # Enhanced summary with column analysis
            summary = {
                'total_rows': len(data_rows),
                'total_columns': len(headers),
                'headers': headers,
                'data_types': self._analyze_data_types(data_rows, headers),
                'column_stats': self._get_column_stats(data_rows, headers),
                'sample_data': data_rows[:5] if data_rows else []  # Show more samples
            }
            
            return {
                'sheet_name': sheet_name,
                'summary': summary,
                'data': data_rows
            }
            
        except Exception as e:
            logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
            return {'error': f"Failed to process sheet {sheet_name}: {str(e)}"}
    
    def _analyze_data_types(self, data_rows: List[Dict], headers: List[str]) -> Dict[str, str]:
        """Enhanced data type analysis"""
        data_types = {}
        
        for header in headers:
            if header.startswith('_'):  # Skip internal fields
                continue
                
            sample_values = [str(row.get(header, "")).strip() for row in data_rows[:20] if row.get(header, "")]
            
            if not sample_values:
                data_types[header] = "empty"
                continue
            
            # Enhanced type detection
            numeric_count = 0
            date_count = 0
            
            for val in sample_values:
                # Check numeric
                if re.match(r'^-?\d+\.?\d*$', val.replace(',', '')):
                    numeric_count += 1
                # Check date patterns
                elif re.match(r'\d{1,4}[-/]\d{1,2}[-/]\d{1,4}', val):
                    date_count += 1
            
            total_samples = len(sample_values)
            if numeric_count > total_samples * 0.7:
                data_types[header] = "numeric"
            elif date_count > total_samples * 0.5:
                data_types[header] = "date"
            else:
                data_types[header] = "text"
                
        return data_types
    
    def _get_column_stats(self, data_rows: List[Dict], headers: List[str]) -> Dict[str, Dict]:
        """Get statistics for each column"""
        stats = {}
        
        for header in headers:
            if header.startswith('_'):
                continue
                
            values = [row.get(header, "") for row in data_rows if row.get(header, "")]
            
            stats[header] = {
                'non_empty_count': len(values),
                'unique_count': len(set(values)) if values else 0,
                'sample_values': list(set(values))[:5] if values else []
            }
            
        return stats
    
    def create_enhanced_summary(self, files_data: List[Dict[str, Any]]) -> str:
        """Create enhanced summary with TF-IDF search capabilities"""
        summary_parts = []
        
        summary_parts.append("=== ENHANCED EXCEL ANALYSIS WITH DYNAMIC DATA RETRIEVAL ===")
        summary_parts.append(f"Total files processed: {len(files_data)}")
        summary_parts.append("📊 Data is indexed and searchable using TF-IDF similarity")
        summary_parts.append("🎯 LLM can request specific rows, columns, and calculations dynamically")
        
        # Build search index
        self.search_engine.build_index(files_data)
        total_rows = len(self.search_engine.documents)
        summary_parts.append(f"🔍 Total searchable rows: {total_rows}")
        
        for file_data in files_data:
            if 'error' in file_data:
                summary_parts.append(f"\n❌ {file_data.get('file_name', 'Unknown')}: {file_data['error']}")
                continue
                
            file_name = file_data['file_name']
            summary_parts.append(f"\n📁 FILE: {file_name}")
            summary_parts.append(f"   Sheets: {file_data['summary']['total_sheets']} ({', '.join(file_data['summary']['sheet_names'])})")
            
            for sheet_name, sheet_data in file_data['sheets'].items():
                if 'error' in sheet_data:
                    summary_parts.append(f"   ❌ Sheet '{sheet_name}': {sheet_data['error']}")
                    continue
                    
                summary = sheet_data['summary']
                summary_parts.append(f"\n   📊 SHEET: {sheet_name}")
                summary_parts.append(f"      Dimensions: {summary['total_rows']} rows × {summary['total_columns']} columns")
                summary_parts.append(f"      Headers: {', '.join(summary['headers'][:10])}")
                
                # Add column statistics
                if 'column_stats' in summary:
                    summary_parts.append("      Column Statistics:")
                    for col, stats in list(summary['column_stats'].items())[:5]:
                        if not col.startswith('_'):
                            summary_parts.append(f"        {col}: {stats['non_empty_count']} values, {stats['unique_count']} unique")
                
                # Enhanced sample data with row numbers
                if summary['sample_data']:
                    summary_parts.append("      Sample rows (with row numbers):")
                    for row in summary['sample_data'][:3]:
                        row_num = row.get('_row_number', '?')
                        clean_row = {k: str(v)[:40] for k, v in row.items() if not k.startswith('_') and v}
                        summary_parts.append(f"        Row {row_num}: {clean_row}")
        
        # Add dynamic retrieval instructions
        summary_parts.append("\n🎯 DYNAMIC DATA RETRIEVAL CAPABILITIES:")
        summary_parts.append("You can request specific data during analysis using:")
        summary_parts.append("- GET_ROW filename sheetname row_number - Get specific row data")
        summary_parts.append("- GET_COLUMN filename sheetname column_name - Get all values from a column")
        summary_parts.append("- GET_STATS filename sheetname column_name - Calculate statistics for numeric column")
        summary_parts.append("- SEARCH query_text - Search for specific data using TF-IDF")
        summary_parts.append("- SUM query_text - Search and sum numeric values")
        summary_parts.append("- Natural language: 'get row 5 from balance sheet', 'calculate total expenses'")
        
        return "\n".join(summary_parts)

class EnhancedGeminiLLM:
    """Enhanced Gemini LLM with dynamic data retrieval capabilities"""
    
    def __init__(self, api_key: str, data_retriever: DataRetriever):
        genai.configure(api_key=api_key)
        self.model = "gemini-1.5-flash"
        self.data_retriever = data_retriever
        
    def process_data_request(self, request: str) -> str:
        """Process LLM data requests and return formatted results"""
        try:
            # Parse common request patterns
            request_lower = request.lower()
            
            # Pattern 1: Get specific row
            if 'row' in request_lower and any(num.isdigit() for num in request.split()):
                # Extract row number
                row_num = None
                for part in request.split():
                    if part.isdigit():
                        row_num = int(part)
                        break
                
                if row_num and st.session_state.files_data:
                    # Use first available file/sheet for demo
                    file_name = st.session_state.files_data[0]['file_name']
                    sheet_name = list(st.session_state.files_data[0]['sheets'].keys())[0]
                    result = self.data_retriever.get_specific_row(file_name, sheet_name, row_num)
                    return self._format_result(result)
            
            # Pattern 2: Get column data
            elif 'column' in request_lower:
                result = self.data_retriever.search_and_aggregate(request, 'list')
                return self._format_result(result)
            
            # Pattern 3: Calculate statistics
            elif any(word in request_lower for word in ['sum', 'total', 'average', 'statistics', 'calculate']):
                result = self.data_retriever.search_and_aggregate(request, 'sum')
                return self._format_result(result)
            
            # Pattern 4: Search and list
            else:
                result = self.data_retriever.search_and_aggregate(request, 'list')
                return self._format_result(result)
                
        except Exception as e:
            return f"Error processing request: {str(e)}"
    
    def _format_result(self, result: Dict[str, Any]) -> str:
        """Format data retrieval results for LLM"""
        if not result.get('success', False):
            return f"❌ Error: {result.get('error', 'Unknown error')}"
        
        formatted = "✅ Data Retrieved Successfully:\n\n"
        
        if 'data' in result:
            if isinstance(result['data'], list):
                formatted += f"📊 Found {len(result['data'])} items:\n"
                for i, item in enumerate(result['data'][:5], 1):  # Show first 5
                    formatted += f"{i}. {item}\n"
                if len(result['data']) > 5:
                    formatted += f"... and {len(result['data']) - 5} more items\n"
            else:
                formatted += f"📊 Data: {result['data']}\n"
        
        if 'statistics' in result:
            stats = result['statistics']
            formatted += f"📈 Statistics:\n"
            formatted += f"  • Count: {stats['count']}\n"
            formatted += f"  • Sum: {stats['sum']:,.2f}\n"
            formatted += f"  • Average: {stats['average']:,.2f}\n"
            formatted += f"  • Min: {stats['min']:,.2f}\n"
            formatted += f"  • Max: {stats['max']:,.2f}\n"
        
        if 'results' in result:
            formatted += f"🔍 Search Results ({result.get('total_found', 0)} total):\n"
            for i, res in enumerate(result['results'][:3], 1):
                formatted += f"{i}. File: {res['file_name']}, Sheet: {res['sheet_name']}, Row: {res['row_index']+1}\n"
                formatted += f"   Score: {res['similarity_score']:.3f}\n"
                # Show a few key-value pairs
                row_preview = {k: v for k, v in list(res['row_data'].items())[:3] if not k.startswith('_') and v}
                formatted += f"   Data: {row_preview}\n\n"
        
        if 'location' in result:
            formatted += f"📍 Location: {result['location']}\n"
            
        return formatted
    
    def analyze_with_dynamic_retrieval(self, excel_summary: str, user_query: str = "") -> str:
        """Enhanced analysis with dynamic data retrieval"""
        try:
            # First, perform initial analysis
            initial_prompt = f"""
You are an expert data analyst with dynamic data retrieval capabilities. 

EXCEL DATA SUMMARY:
{excel_summary}

USER QUERY: {user_query if user_query else "Provide comprehensive analysis"}

Based on this summary, you can request specific data using these commands:
- "GET_ROW file_name sheet_name row_number" - Get specific row data
- "GET_COLUMN file_name sheet_name column_name" - Get all values from a column  
- "GET_STATS file_name sheet_name column_name" - Calculate statistics for numeric column
- "SEARCH query_text" - Search for specific data using TF-IDF
- "SUM query_text" - Search and sum numeric values

Please provide your initial analysis, and if you need specific data to give better insights, format your requests clearly using the commands above.

Focus on:
1. Initial insights from the summary
2. Specific data requests you need to provide deeper analysis
3. What patterns you want to investigate further
"""

            model = genai.GenerativeModel(self.model)
            initial_response = model.generate_content(initial_prompt)
            
            # Parse response for data requests
            data_requests = self._parse_data_requests(initial_response.text)
            
            # Process data requests
            retrieved_data = ""
            if data_requests:
                retrieved_data = "\n🔍 RETRIEVED DATA:\n"
                for request in data_requests[:3]:  # Limit to 3 requests to avoid long responses
                    result = self.process_data_request(request)
                    retrieved_data += f"\nRequest: {request}\n{result}\n"
            
            # Generate final analysis with retrieved data
            if retrieved_data:
                final_prompt = f"""
Based on your initial analysis and the retrieved data below, provide a comprehensive final analysis:

INITIAL ANALYSIS:
{initial_response.text}

{retrieved_data}

Now provide:
1. Updated insights based on the retrieved data
2. Specific findings with exact numbers and locations
3. Data quality observations with examples
4. Actionable recommendations
5. Further investigation suggestions
"""
                final_response = model.generate_content(final_prompt)
                
                return f"{initial_response.text}\n\n{retrieved_data}\n\n🔄 ENHANCED ANALYSIS:\n{final_response.text}"
            else:
                return initial_response.text
                
        except Exception as e:
            logger.error(f"Error in dynamic analysis: {str(e)}")
            return f"Error in analysis: {str(e)}"
    
    def _parse_data_requests(self, text: str) -> List[str]:
        """Parse data requests from LLM response"""
        requests = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            if any(cmd in line.upper() for cmd in ['GET_ROW', 'GET_COLUMN', 'GET_STATS', 'SEARCH', 'SUM']):
                requests.append(line)
        
        return requests

# Initialize session state
if 'files_data' not in st.session_state:
    st.session_state.files_data = []
if 'excel_summary' not in st.session_state:
    st.session_state.excel_summary = ""
if 'llm_analysis' not in st.session_state:
    st.session_state.llm_analysis = ""
if 'search_results' not in st.session_state:
    st.session_state.search_results = []
if 'excel_processor' not in st.session_state:
    st.session_state.excel_processor = ExcelProcessor()
if 'data_retriever' not in st.session_state:
    st.session_state.data_retriever = None

# Main App
def main():
    st.title("📊 Excel AI Analyzer with Dynamic Data Retrieval")
    st.markdown("Upload Excel files, build searchable index, and get AI insights with intelligent data retrieval!")
    
    # Sidebar for configuration
    with st.sidebar:
        st.header("🔧 Configuration")
        
        # API Key input
        api_key = st.text_input(
            "🔑 Gemini API Key",
            type="password",
            help="Enter your Google Gemini API key"
        )
        
        if api_key:
            st.success("✅ API Key set")
        else:
            st.warning("⚠️ Enter Gemini API key")
        
        st.divider()
        
        # Processing settings
        st.header("⚙️ Processing Settings")
        max_rows = st.slider("Max rows per sheet", 50, 500, 100)
        max_chars = st.slider("Max characters per cell", 200, 1000, 500)
        
        # Update processor settings
        st.session_state.excel_processor.max_rows_per_sheet = max_rows
        st.session_state.excel_processor.max_chars_per_cell = max_chars
        
        st.divider()
        
        # TF-IDF Search Section
        st.header("🔍 TF-IDF Search")
        
        if st.session_state.files_data:
            search_query = st.text_input(
                "Search your data:",
                placeholder="e.g., sales revenue Q1, customer data, highest values"
            )
            
            search_button = st.button("🔍 Search", type="secondary")
            
            if search_button and search_query:
                with st.spinner("Searching..."):
                    results = st.session_state.excel_processor.search_engine.search(search_query, top_k=10)
                    st.session_state.search_results = results
                    
                if results:
                    st.success(f"Found {len(results)} results")
                    
                    # Show top results in sidebar
                    for i, result in enumerate(results[:3], 1):
                        with st.expander(f"Result {i} (Score: {result['similarity_score']:.3f})"):
                            st.write(f"**File:** {result['file_name']}")
                            st.write(f"**Sheet:** {result['sheet_name']}")
                            st.write(f"**Row:** {result['row_index'] + 1}")
                            st.json(result['row_data'], expanded=False)
                else:
                    st.warning("No results found")
            
            # Advanced Data Retrieval Section
            st.subheader("🎯 Advanced Data Retrieval")
            
            if st.session_state.data_retriever:
                # File selector
                file_options = [f['file_name'] for f in st.session_state.files_data if 'error' not in f]
                selected_file = st.selectbox("Select File:", file_options, key="sidebar_file")
                
                if selected_file:
                    # Sheet selector
                    file_data = next(f for f in st.session_state.files_data if f.get('file_name') == selected_file)
                    sheet_options = list(file_data['sheets'].keys())
                    selected_sheet = st.selectbox("Select Sheet:", sheet_options, key="sidebar_sheet")
                    
                    # Quick data retrieval buttons
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("📊 Get Sheet Summary", key="get_summary"):
                            result = st.session_state.data_retriever.get_sheet_summary(selected_file, selected_sheet)
                            st.json(result, expanded=False)
                    
                    with col2:
                        row_number = st.number_input("Row #:", min_value=1, value=1, key="row_input")
                        if st.button("📄 Get Row", key="get_row"):
                            result = st.session_state.data_retriever.get_specific_row(selected_file, selected_sheet, row_number)
                            st.json(result, expanded=False)
                    
                    # Column analysis
                    if selected_sheet and selected_file:
                        file_data = next(f for f in st.session_state.files_data if f.get('file_name') == selected_file)
                        if selected_sheet in file_data['sheets']:
                            headers = file_data['sheets'][selected_sheet]['summary'].get('headers', [])
                            if headers:
                                selected_column = st.selectbox("Select Column:", headers, key="sidebar_column")
                                
                                col1, col2 = st.columns(2)
                                with col1:
                                    if st.button("📈 Get Column", key="get_column"):
                                        result = st.session_state.data_retriever.get_column_values(selected_file, selected_sheet, selected_column)
                                        st.json(result, expanded=False)
                                
                                with col2:
                                    if st.button("🔢 Calculate Stats", key="get_stats"):
                                        result = st.session_state.data_retriever.calculate_column_statistics(selected_file, selected_sheet, selected_column)
                                        st.json(result, expanded=False)
        else:
            st.info("Process files first to enable search")
        
        st.divider()
        
        # Clear button
        if st.button("🗑️ Clear All Data", type="secondary"):
            for key in ['files_data', 'excel_summary', 'llm_analysis', 'search_results']:
                st.session_state[key] = [] if 'results' in key or 'data' in key else ""
            st.session_state.excel_processor = ExcelProcessor()
            st.session_state.data_retriever = None
            st.rerun()
    
    # Main content area
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.header("📁 Upload Excel Files")
        
        uploaded_files = st.file_uploader(
            "Choose Excel files",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            help="Upload multiple Excel files (.xlsx or .xls format)"
        )
        
        if uploaded_files:
            st.success(f"📄 {len(uploaded_files)} file(s) uploaded")
            for file in uploaded_files:
                st.write(f"• {file.name} ({file.size / 1024 / 1024:.2f} MB)")
    
    with col2:
        st.header("❓ Your Question (Optional)")
        
        user_query = st.text_area(
            "Ask about your Excel data",
            placeholder="e.g., What are the sales trends? Find customers with highest revenue. Show me data quality issues.",
            height=150
        )
    
    # Processing section
    st.divider()
    
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    
    with col1:
        process_button = st.button(
            "🚀 Process & Index",
            type="primary",
            disabled=not uploaded_files,
            help="Process Excel files and build TF-IDF search index"
        )
    
    with col2:
        analyze_button = st.button(
            "🤖 AI Analysis",
            type="primary",
            disabled=not (st.session_state.excel_summary and api_key),
            help="Get AI insights with dynamic data retrieval"
        )
    
    with col3:
        if st.session_state.search_results:
            st.metric("Search Results", len(st.session_state.search_results))
        else:
            st.metric("Search Results", 0)
    
    with col4:
        if st.session_state.excel_processor.search_engine.documents:
            st.metric("Indexed Rows", len(st.session_state.excel_processor.search_engine.documents))
        else:
            st.metric("Indexed Rows", 0)
    
    # Process files
    if process_button and uploaded_files:
        with st.spinner("🔄 Processing files and building search index..."):
            files_data = []
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            for i, uploaded_file in enumerate(uploaded_files):
                status_text.text(f"Processing {uploaded_file.name}...")
                progress_bar.progress((i + 1) / len(uploaded_files))
                
                file_content = uploaded_file.read()
                file_data = st.session_state.excel_processor.read_excel_file(file_content, uploaded_file.name)
                files_data.append(file_data)
            
            st.session_state.files_data = files_data
            st.session_state.excel_summary = st.session_state.excel_processor.create_enhanced_summary(files_data)
            
            # Initialize data retriever
            st.session_state.data_retriever = DataRetriever(st.session_state.excel_processor)
            
            progress_bar.progress(1.0)
            status_text.text("✅ Processing complete! TF-IDF index and data retriever ready.")
            
            st.success(f"Successfully processed {len(files_data)} files with {len(st.session_state.excel_processor.search_engine.documents)} searchable rows!")
    
    # AI Analysis with Dynamic Retrieval
    if analyze_button and st.session_state.excel_summary and api_key:
        with st.spinner("🤖 Analyzing with AI and dynamic data retrieval..."):
            try:
                if not st.session_state.data_retriever:
                    st.session_state.data_retriever = DataRetriever(st.session_state.excel_processor)
                
                enhanced_llm = EnhancedGeminiLLM(api_key, st.session_state.data_retriever)
                
                analysis = enhanced_llm.analyze_with_dynamic_retrieval(
                    st.session_state.excel_summary, 
                    user_query
                )
                st.session_state.llm_analysis = analysis
                st.success("✅ AI analysis with dynamic data retrieval complete!")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
    
    # Results section
    if st.session_state.excel_summary or st.session_state.llm_analysis or st.session_state.search_results:
        st.divider()
        st.header("📊 Results")
        
        tab1, tab2, tab3, tab4, tab5 = st.tabs(["📋 Data Summary", "🤖 AI Analysis", "🔍 Search Results", "🎯 Data Retrieval", "📈 Row Browser"])
        
        with tab1:
            if st.session_state.excel_summary:
                st.subheader("📊 Enhanced Data Summary with Dynamic Retrieval")
                st.text_area(
                    "Processed data summary",
                    value=st.session_state.excel_summary,
                    height=400,
                    disabled=True
                )
                
                st.download_button(
                    label="📥 Download Summary",
                    data=st.session_state.excel_summary,
                    file_name="excel_summary_with_dynamic_retrieval.txt",
                    mime="text/plain"
                )
            else:
                st.info("👆 Process Excel files first")
        
        with tab2:
            if st.session_state.llm_analysis:
                st.subheader("🤖 AI Analysis with Dynamic Data Retrieval")
                st.markdown(st.session_state.llm_analysis)
                
                st.download_button(
                    label="📥 Download Analysis",
                    data=st.session_state.llm_analysis,
                    file_name="ai_analysis_with_retrieval.txt",
                    mime="text/plain"
                )
            else:
                st.info("👆 Run AI analysis to see insights")
        
        with tab3:
            if st.session_state.search_results:
                st.subheader("🔍 TF-IDF Search Results")
                
                for i, result in enumerate(st.session_state.search_results, 1):
                    with st.expander(f"📄 Result {i} - Similarity: {result['similarity_score']:.3f}"):
                        col1, col2, col3 = st.columns([1, 1, 1])
                        col1.metric("File", result['file_name'])
                        col2.metric("Sheet", result['sheet_name'])
                        col3.metric("Row Number", result['row_index'] + 1)
                        
                        st.subheader("Row Data:")
                        # Create a nice table view
                        row_df = pd.DataFrame([result['row_data']])
                        st.dataframe(row_df, use_container_width=True)
                        
                        st.subheader("Matched Text:")
                        st.text(result['matched_text'])
            else:
                st.info("👆 Use the search feature in the sidebar")
        
        with tab4:
            if st.session_state.data_retriever:
                st.subheader("🎯 Interactive Data Retrieval")
                st.markdown("Test the LLM's data retrieval capabilities manually:")
                
                # Manual data request interface
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    manual_request = st.text_input(
                        "Enter data request:",
                        placeholder="e.g., get row 5 from balance sheet, sum all expenses, find accommodation costs",
                        key="manual_request"
                    )
                
                with col2:
                    if st.button("📊 Execute Request", key="execute_request"):
                        if manual_request and st.session_state.data_retriever:
                            with st.spinner("Processing request..."):
                                enhanced_llm = EnhancedGeminiLLM("dummy", st.session_state.data_retriever)
                                result = enhanced_llm.process_data_request(manual_request)
                                st.markdown("**Result:**")
                                st.markdown(result)
                
                st.divider()
                
                # Predefined quick actions
                st.subheader("🚀 Quick Actions")
                
                if st.session_state.files_data:
                    file_options = [f['file_name'] for f in st.session_state.files_data if 'error' not in f]
                    
                    if file_options:
                        selected_file = st.selectbox("Choose file for quick actions:", file_options, key="quick_file")
                        file_data = next(f for f in st.session_state.files_data if f.get('file_name') == selected_file)
                        sheet_options = list(file_data['sheets'].keys())
                        
                        selected_sheet = st.selectbox("Choose sheet:", sheet_options, key="quick_sheet")
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            if st.button("📊 Analyze Sheet", key="analyze_sheet"):
                                result = st.session_state.data_retriever.get_sheet_summary(selected_file, selected_sheet)
                                if result['success']:
                                    st.json(result['summary'])
                                    if result.get('sample_data'):
                                        st.subheader("Sample Data:")
                                        df = pd.DataFrame(result['sample_data'])
                                        st.dataframe(df)
                                else:
                                    st.error(result['error'])
                        
                        with col2:
                            row_num = st.number_input("Row number:", min_value=1, value=1, key="quick_row")
                            if st.button("📄 Get Row Data", key="get_row_data"):
                                result = st.session_state.data_retriever.get_specific_row(selected_file, selected_sheet, row_num)
                                if result['success']:
                                    st.json(result['data'])
                                else:
                                    st.error(result['error'])
                        
                        with col3:
                            # Column selector
                            if selected_sheet in file_data['sheets']:
                                headers = file_data['sheets'][selected_sheet]['summary'].get('headers', [])
                                if headers:
                                    selected_col = st.selectbox("Column:", headers, key="quick_col")
                                    if st.button("📈 Column Stats", key="col_stats"):
                                        result = st.session_state.data_retriever.calculate_column_statistics(selected_file, selected_sheet, selected_col)
                                        if result['success']:
                                            st.json(result['statistics'])
                                        else:
                                            st.error(result['error'])
                
                st.divider()
                
                # Command examples
                st.subheader("💡 Command Examples")
                st.markdown("""
                **LLM can request data using these patterns:**
                
                - `GET_ROW filename sheetname 5` - Get specific row
                - `GET_COLUMN filename sheetname Amount` - Get all values from Amount column
                - `GET_STATS filename sheetname Revenue` - Calculate statistics for Revenue column
                - `SEARCH accommodation expenses` - Search for accommodation-related data
                - `SUM total expenses august` - Search and sum expense values
                
                **Natural language examples:**
                - "Show me row 10 from the balance sheet"
                - "Calculate total revenue from all sheets"
                - "Find all accommodation expenses"
                - "Get statistics for the Amount column"
                """)
            else:
                st.info("Process Excel files first to enable data retrieval")
        
        with tab5:
            if st.session_state.files_data:
                st.subheader("📈 Row Browser - View All Data")
                
                # File and sheet selector
                file_options = [f['file_name'] for f in st.session_state.files_data if 'error' not in f]
                if file_options:
                    selected_file = st.selectbox("Select File:", file_options)
                    
                    # Get selected file data
                    file_data = next(f for f in st.session_state.files_data if f.get('file_name') == selected_file)
                    sheet_options = list(file_data['sheets'].keys())
                    
                    if sheet_options:
                        selected_sheet = st.selectbox("Select Sheet:", sheet_options)
                        
                        # Display sheet data
                        sheet_data = file_data['sheets'][selected_sheet]
                        if 'error' not in sheet_data and sheet_data['data']:
                            st.subheader(f"📊 {selected_sheet} - All Rows")
                            
                            # Convert to DataFrame for better display
                            df = pd.DataFrame(sheet_data['data'])
                            
                            # Remove internal columns
                            display_df = df.drop(columns=[col for col in df.columns if col.startswith('_')])
                            
                            # Add row numbers
                            display_df.insert(0, 'Row #', df['_row_number'] if '_row_number' in df.columns else range(1, len(df) + 1))
                            
                            st.dataframe(display_df, use_container_width=True, height=400)
                            
                            # Download option
                            csv = display_df.to_csv(index=False)
                            st.download_button(
                                label="📥 Download as CSV",
                                data=csv,
                                file_name=f"{selected_file}_{selected_sheet}.csv",
                                mime="text/csv"
                            )
                        else:
                            st.warning("No data to display for this sheet")
            else:
                st.info("👆 Process Excel files first")

if __name__ == "__main__":
    main()
